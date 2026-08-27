from __future__ import annotations

import hashlib
import json
import os
import shutil
import sqlite3
from contextlib import closing
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from backend import MIGRATION_VERSION
from backend.db import connect, initialize_schema
from backend.repository import company_api_to_excel, utc_now
from excel_tools import read_company_rows


@dataclass
class MigrationPlan:
    report: dict[str, Any]
    companies: list[dict[str, Any]]
    jobs: list[dict[str, Any]]
    applications: list[dict[str, Any]]
    candidates: list[dict[str, Any]]
    utility_runs: list[dict[str, Any]]


def build_migration_plan(project_root: Path, database_path: Path) -> MigrationPlan:
    project_root = Path(project_root).resolve()
    database_path = Path(database_path).resolve()
    source_paths = detected_source_paths(project_root, database_path)
    source_entries = [describe_source(path, project_root) for path in source_paths]

    company_rows = read_company_rows(project_root / "data" / "master.xlsx")
    raw_jobs = read_json(project_root / "data" / "jobs.json", [])
    raw_applications = read_json(project_root / "data" / "applications.json", [])
    raw_candidates = read_json(project_root / "logs" / "rejected_job_candidates.json", [])
    diagnostics = [
        path for path in (
            project_root / "logs" / "job_board_discovery_audit.json",
            project_root / "logs" / "job_collection_diagnostics.json",
            project_root / "logs" / "website_audit.json",
        ) if path.exists()
    ]

    actions: dict[str, list[dict[str, Any]]] = {name: [] for name in (
        "companies", "jobs", "applications", "raw_job_candidates", "utility_runs"
    )}
    findings: dict[str, list[dict[str, Any]]] = {
        "duplicates": [], "invalidRecords": [], "stableIdConflicts": [], "relationshipRepairs": []
    }
    existing_ids = load_existing_ids(database_path)

    companies: list[dict[str, Any]] = []
    company_ids: set[str] = set()
    companies_by_name: dict[str, list[str]] = defaultdict(list)
    for index, row in enumerate(company_rows):
        company = excel_company_to_api(row)
        company_id = company["id"]
        if not company_id or not company["name"]:
            reason = "missing stable company ID" if not company_id else "missing company name"
            record = {"sourceIndex": index, "id": company_id, "reason": reason}
            findings["invalidRecords"].append({"table": "companies", **record})
            actions["companies"].append({"action": "skip", **record})
            continue
        if company_id in company_ids:
            record = {"table": "companies", "id": company_id, "sourceIndex": index, "reason": "duplicate stable ID"}
            findings["duplicates"].append(record)
            findings["stableIdConflicts"].append(record)
            actions["companies"].append({"action": "skip", **record})
            continue
        company_ids.add(company_id)
        companies_by_name[company["name"].casefold()].append(company_id)
        companies.append(company)
        actions["companies"].append(action_for_id("companies", company_id, existing_ids))

    jobs: list[dict[str, Any]] = []
    used_job_ids: set[str] = set()
    legacy_job_map: dict[str, list[str]] = defaultdict(list)
    for index, raw in enumerate(raw_jobs if isinstance(raw_jobs, list) else []):
        if not isinstance(raw, dict):
            record = {"table": "jobs", "sourceIndex": index, "reason": "record is not an object"}
            findings["invalidRecords"].append(record)
            actions["jobs"].append({"action": "skip", **record})
            continue
        legacy_id = str(raw.get("id") or "").strip()
        if not legacy_id or not str(raw.get("title") or "").strip():
            reason = "missing job ID" if not legacy_id else "missing job title"
            record = {"table": "jobs", "id": legacy_id, "sourceIndex": index, "reason": reason}
            findings["invalidRecords"].append(record)
            actions["jobs"].append({"action": "skip", **record})
            continue
        job_id = legacy_id
        if job_id in used_job_ids:
            job_id = repaired_job_id(legacy_id, raw, used_job_ids)
            conflict = {
                "table": "jobs", "sourceIndex": index, "originalId": legacy_id,
                "resolvedId": job_id, "reason": "duplicate legacy job ID repaired deterministically",
            }
            findings["duplicates"].append(conflict)
            findings["stableIdConflicts"].append(conflict)
            actions["jobs"].append({"action": "update-id", **conflict})
        else:
            actions["jobs"].append(action_for_id("jobs", job_id, existing_ids))
        used_job_ids.add(job_id)
        legacy_job_map[legacy_id].append(job_id)
        job = dict(raw)
        job["id"] = job_id
        job["legacyId"] = legacy_id
        source_company_id = str(raw.get("companyId") or "")
        if source_company_id not in company_ids:
            matches = companies_by_name.get(str(raw.get("companyName") or "").casefold(), [])
            if len(matches) == 1:
                job["companyId"] = matches[0]
                findings["relationshipRepairs"].append({
                    "table": "jobs", "id": job_id, "fromCompanyId": source_company_id,
                    "toCompanyId": matches[0], "reason": "resolved by unique company name",
                })
            else:
                job["companyId"] = ""
                findings["invalidRecords"].append({
                    "table": "jobs", "id": job_id, "reason": "company relationship unresolved; preserved with null foreign key",
                })
        jobs.append(job)

    applications = normalize_applications(raw_applications, legacy_job_map, used_job_ids, actions, findings, existing_ids)
    candidates = normalize_candidates(raw_candidates, companies_by_name, company_ids, actions, findings, existing_ids)
    utility_runs = normalize_utility_runs(diagnostics, actions, existing_ids)

    counts = {
        "source": {
            "companies": len(company_rows),
            "jobs": len(raw_jobs) if isinstance(raw_jobs, list) else 0,
            "applications": len(raw_applications) if isinstance(raw_applications, (list, dict)) else 0,
            "rawJobCandidates": len(raw_candidates) if isinstance(raw_candidates, list) else 0,
        },
        "importable": {
            "companies": len(companies), "jobs": len(jobs), "applications": len(applications),
            "rawJobCandidates": len(candidates), "utilityRuns": len(utility_runs),
            "resumes": 0, "resumeFitResults": 0, "settings": 0,
        },
        "findings": {key: len(value) for key, value in findings.items()},
        "actions": {
            table: dict(Counter(item["action"] for item in table_actions))
            for table, table_actions in actions.items()
        },
    }
    report = {
        "migrationVersion": MIGRATION_VERSION,
        "mode": "preview",
        "generatedAt": utc_now(),
        "projectRoot": str(project_root),
        "database": str(database_path),
        "databaseExists": database_path.exists(),
        "sourceFilesDetected": source_entries,
        "counts": counts,
        "findings": findings,
        "actions": actions,
    }
    return MigrationPlan(report, companies, jobs, applications, candidates, utility_runs)


def apply_migration(project_root: Path, database_path: Path) -> dict[str, Any]:
    project_root = Path(project_root).resolve()
    database_path = Path(database_path).resolve()
    plan = build_migration_plan(project_root, database_path)
    started_at = utc_now()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = project_root / "data" / "exports" / f"pre_sqlite_migration_{stamp}"
    manifest_path, manifest_hash = create_backup(project_root, database_path, backup_dir)
    output_dir = project_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    report_json = output_dir / f"sqlite_migration_report_{stamp}.json"
    report_xlsx = output_dir / f"sqlite_migration_report_{stamp}.xlsx"
    temporary_database = database_path.with_name(f".{database_path.name}.migrating-{uuid4().hex}")

    try:
        if database_path.exists():
            with closing(connect(database_path, readonly=True)) as source, closing(sqlite3.connect(temporary_database)) as destination:
                source.backup(destination)
        connection = connect(temporary_database)
        try:
            initialize_schema(connection)
            connection.execute("BEGIN IMMEDIATE")
            apply_plan(connection, plan)
            run_crud_cascade_probe(connection)
            completed_at = utc_now()
            import_id = f"migration-{stamp}-{uuid4().hex[:8]}"
            connection.execute(
                """INSERT INTO import_history (id, migration_version, source_manifest_path,
                source_manifest_sha256, report_json_path, started_at, completed_at, status, counts_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'completed', ?)""",
                (
                    import_id, MIGRATION_VERSION, str(manifest_path), manifest_hash, str(report_json),
                    started_at, completed_at, json.dumps(plan.report["counts"], sort_keys=True),
                ),
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()

        validation = validate_database(temporary_database, plan)
        if not validation["passed"]:
            raise RuntimeError("SQLite migration validation failed: " + "; ".join(validation["errors"]))

        final_report = dict(plan.report)
        final_report.update({
            "mode": "apply", "status": "completed", "startedAt": started_at,
            "completedAt": utc_now(), "backupDirectory": str(backup_dir),
            "manifest": str(manifest_path), "reportJson": str(report_json),
            "reportXlsx": str(report_xlsx), "validation": validation,
        })
        write_reports(final_report, report_json, report_xlsx)
        database_path.parent.mkdir(parents=True, exist_ok=True)
        os.replace(temporary_database, database_path)
        cleanup_sqlite_sidecars(temporary_database)
        return final_report
    except Exception:
        temporary_database.unlink(missing_ok=True)
        cleanup_sqlite_sidecars(temporary_database)
        report_json.unlink(missing_ok=True)
        report_xlsx.unlink(missing_ok=True)
        raise


def apply_plan(connection: sqlite3.Connection, plan: MigrationPlan) -> None:
    now = utc_now()
    for company in plan.companies:
        connection.execute(
            """INSERT INTO companies (id,name,industry,city,state,country,known_website,official_website,
            website_discovery_method,website_candidate_urls,website_verification_notes,website_verified,
            careers_page_url,job_board_url,job_board_discovery_method,jobs_rss_feed_url,job_platform,
            feed_found,search_status,confidence,last_checked,notes,founded_year,total_assets,
            assets_as_of_date,company_info_last_checked,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET name=excluded.name,industry=excluded.industry,city=excluded.city,
            state=excluded.state,country=excluded.country,known_website=excluded.known_website,
            official_website=excluded.official_website,website_discovery_method=excluded.website_discovery_method,
            website_candidate_urls=excluded.website_candidate_urls,website_verification_notes=excluded.website_verification_notes,
            website_verified=excluded.website_verified,careers_page_url=excluded.careers_page_url,
            job_board_url=excluded.job_board_url,job_board_discovery_method=excluded.job_board_discovery_method,
            jobs_rss_feed_url=excluded.jobs_rss_feed_url,job_platform=excluded.job_platform,
            feed_found=excluded.feed_found,search_status=excluded.search_status,confidence=excluded.confidence,
            last_checked=excluded.last_checked,notes=excluded.notes,founded_year=excluded.founded_year,
            total_assets=excluded.total_assets,assets_as_of_date=excluded.assets_as_of_date,
            company_info_last_checked=excluded.company_info_last_checked,updated_at=excluded.updated_at""",
            company_values(company, now),
        )
    for job in plan.jobs:
        connection.execute(
            """INSERT INTO jobs (id,legacy_id,company_id,company_name,title,location,work_type,pay_min,pay_max,
            pay_text,pay_period,pay_currency,posted_date,source_url,job_platform,description,description_snippet,
            collected_at,status,role_type,role_type_reason,raw_data_json,first_seen_at,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET legacy_id=excluded.legacy_id,company_id=excluded.company_id,
            company_name=excluded.company_name,title=excluded.title,location=excluded.location,
            work_type=excluded.work_type,pay_min=excluded.pay_min,pay_max=excluded.pay_max,pay_text=excluded.pay_text,
            pay_period=excluded.pay_period,pay_currency=excluded.pay_currency,posted_date=excluded.posted_date,
            source_url=excluded.source_url,job_platform=excluded.job_platform,description=excluded.description,
            description_snippet=excluded.description_snippet,collected_at=excluded.collected_at,status=excluded.status,
            role_type=excluded.role_type,role_type_reason=excluded.role_type_reason,
            raw_data_json=excluded.raw_data_json,updated_at=excluded.updated_at""",
            job_values(job, now),
        )
    for application in plan.applications:
        connection.execute(
            """INSERT INTO applications (job_id,applied,application_status,date_applied,follow_up_date,notes,
            not_interested,payload_json,archived_at,updated_at) VALUES (?,?,?,?,?,?,?,?,NULL,?)
            ON CONFLICT(job_id) DO UPDATE SET applied=excluded.applied,application_status=excluded.application_status,
            date_applied=excluded.date_applied,follow_up_date=excluded.follow_up_date,notes=excluded.notes,
            not_interested=excluded.not_interested,payload_json=excluded.payload_json,updated_at=excluded.updated_at""",
            application_values(application, now),
        )
    for candidate in plan.candidates:
        connection.execute(
            """INSERT INTO raw_job_candidates (id,company_id,job_id,company_name,candidate_text,candidate_href,
            rejection_reason,payload_json,imported_at) VALUES (?,?,?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET company_id=excluded.company_id,job_id=excluded.job_id,
            company_name=excluded.company_name,candidate_text=excluded.candidate_text,
            candidate_href=excluded.candidate_href,rejection_reason=excluded.rejection_reason,
            payload_json=excluded.payload_json""",
            (
                candidate["id"], candidate.get("companyId") or None, candidate.get("jobId") or None,
                candidate.get("companyName", ""), candidate.get("candidateText", ""), candidate.get("candidateHref", ""),
                candidate.get("rejectionReason", ""), json.dumps(candidate["payload"], sort_keys=True), now,
            ),
        )
    for run in plan.utility_runs:
        connection.execute(
            """INSERT INTO utility_runs (id,utility_name,status,started_at,completed_at,payload_json,created_at)
            VALUES (?,?,?,?,?,?,?) ON CONFLICT(id) DO UPDATE SET status=excluded.status,
            payload_json=excluded.payload_json""",
            (run["id"], run["utilityName"], run.get("status", ""), "", "", json.dumps(run["payload"], sort_keys=True), now),
        )


def validate_database(database_path: Path, plan: MigrationPlan) -> dict[str, Any]:
    errors: list[str] = []
    expected = {
        "companies": {item["id"] for item in plan.companies},
        "jobs": {item["id"] for item in plan.jobs},
        "applications": {item["jobId"] for item in plan.applications},
        "raw_job_candidates": {item["id"] for item in plan.candidates},
        "utility_runs": {item["id"] for item in plan.utility_runs},
    }
    actual_counts: dict[str, int] = {}
    with closing(connect(database_path, readonly=True)) as connection:
        foreign_key_errors = connection.execute("PRAGMA foreign_key_check").fetchall()
        if foreign_key_errors:
            errors.append(f"foreign key check returned {len(foreign_key_errors)} error(s)")
        key_columns = {"companies": "id", "jobs": "id", "applications": "job_id", "raw_job_candidates": "id", "utility_runs": "id"}
        for table, expected_ids in expected.items():
            actual_ids = {row[0] for row in connection.execute(f"SELECT {key_columns[table]} FROM {table}")}
            actual_counts[table] = len(actual_ids)
            missing = expected_ids - actual_ids
            if missing:
                errors.append(f"{table}: {len(missing)} imported IDs missing")
        table_names = {row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        required = {"companies", "jobs", "raw_job_candidates", "applications", "resumes", "resume_fit_results", "settings", "utility_runs", "import_history"}
        missing_tables = required - table_names
        if missing_tables:
            errors.append("missing required tables: " + ", ".join(sorted(missing_tables)))
    return {"passed": not errors, "errors": errors, "sqliteCounts": actual_counts}


def run_crud_cascade_probe(connection: sqlite3.Connection) -> None:
    company_id = f"company-migration-probe-{uuid4().hex}"
    job_id = f"job-migration-probe-{uuid4().hex}"
    now = utc_now()
    connection.execute("SAVEPOINT crud_probe")
    try:
        probe = {
            "id": company_id, "name": "Migration CRUD Probe", "industry": "Financial Services",
            "city": "", "state": "", "country": "United States", "knownWebsite": "https://probe.invalid",
            "officialWebsite": "https://probe.invalid", "websiteDiscoveryMethod": "Not Found",
            "websiteCandidateUrls": "", "websiteVerificationNotes": "", "websiteVerified": False,
            "careersPageUrl": "", "jobBoardUrl": "", "jobBoardDiscoveryMethod": "Not Found",
            "jobsRssFeedUrl": "", "jobPlatform": "", "feedFound": False, "searchStatus": "Needs Review",
            "confidence": 0, "lastChecked": "", "notes": "",
        }
        connection.execute(
            """INSERT INTO companies (id,name,industry,city,state,country,known_website,official_website,
            website_discovery_method,website_candidate_urls,website_verification_notes,website_verified,
            careers_page_url,job_board_url,job_board_discovery_method,jobs_rss_feed_url,job_platform,
            feed_found,search_status,confidence,last_checked,notes,founded_year,total_assets,
            assets_as_of_date,company_info_last_checked,created_at,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            company_values(probe, now),
        )
        connection.execute("UPDATE companies SET official_website=?, job_board_discovery_method='Manual Re-verification Required', search_status='Needs Review' WHERE id=?", ("https://edited.invalid", company_id))
        if connection.execute("SELECT job_board_discovery_method FROM companies WHERE id=?", (company_id,)).fetchone()[0] != "Manual Re-verification Required":
            raise RuntimeError("CRUD probe did not retain manual re-verification flag")
        probe_job = {"id": job_id, "legacyId": job_id, "companyId": company_id, "companyName": probe["name"], "title": "Probe Job"}
        connection.execute("INSERT INTO jobs VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", job_values(probe_job, now))
        connection.execute("INSERT INTO applications (job_id,payload_json,updated_at) VALUES (?, '{}', ?)", (job_id, now))
        connection.execute("INSERT INTO raw_job_candidates (id,company_id,company_name,payload_json,imported_at) VALUES (?,?,?,'{}',?)", (f"candidate-{uuid4().hex}", company_id, probe["name"], now))
        connection.execute("DELETE FROM companies WHERE id=?", (company_id,))
        if connection.execute("SELECT 1 FROM jobs WHERE id=?", (job_id,)).fetchone():
            raise RuntimeError("CRUD probe job cascade failed")
        if connection.execute("SELECT 1 FROM applications WHERE job_id=?", (job_id,)).fetchone():
            raise RuntimeError("CRUD probe application cascade failed")
        if connection.execute("SELECT 1 FROM raw_job_candidates WHERE company_id=?", (company_id,)).fetchone():
            raise RuntimeError("CRUD probe candidate cascade failed")
    finally:
        connection.execute("ROLLBACK TO crud_probe")
        connection.execute("RELEASE crud_probe")


def create_backup(project_root: Path, database_path: Path, backup_dir: Path) -> tuple[Path, str]:
    backup_dir.mkdir(parents=True, exist_ok=False)
    entries: list[dict[str, Any]] = []
    for source in detected_source_paths(project_root, database_path):
        relative = source.relative_to(project_root)
        destination = backup_dir / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
        entry = describe_source(source, project_root)
        entry["backupPath"] = str(destination.relative_to(backup_dir))
        entries.append(entry)
    manifest = {
        "migrationVersion": MIGRATION_VERSION, "createdAt": utc_now(),
        "projectRoot": str(project_root), "files": entries,
    }
    manifest_path = backup_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest_path, sha256_file(manifest_path)


def detected_source_paths(project_root: Path, database_path: Path) -> list[Path]:
    candidates = [
        project_root / "data" / "master.xlsx", project_root / "data" / "companies.json",
        project_root / "data" / "jobs.json", project_root / "data" / "applications.json",
        project_root / "frontend" / "public" / "data" / "companies.json",
        project_root / "frontend" / "public" / "data" / "jobs.json",
        project_root / "logs" / "rejected_job_candidates.json",
        project_root / "logs" / "job_board_discovery_audit.json",
        project_root / "logs" / "job_collection_diagnostics.json",
        project_root / "logs" / "website_audit.json",
        project_root / "output" / "jobs_snapshot.xlsx",
        project_root / "output" / "rejected_job_candidates.xlsx",
        project_root / "output" / "job_board_discovery_audit.xlsx",
        project_root / "output" / "job_collection_diagnostics.xlsx",
        project_root / "output" / "website_audit.xlsx",
        database_path,
    ]
    return [path.resolve() for path in candidates if path.exists()]


def describe_source(path: Path, project_root: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "path": str(path.relative_to(project_root)) if path.is_relative_to(project_root) else str(path),
        "sizeBytes": stat.st_size, "sha256": sha256_file(path), "rowCount": source_row_count(path),
        "modifiedAt": datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat(),
    }


def source_row_count(path: Path) -> int | None:
    try:
        if path.suffix.lower() == ".json":
            payload = read_json(path, None)
            return len(payload) if isinstance(payload, (list, dict)) else None
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                return max(workbook.active.max_row - 1, 0)
            finally:
                workbook.close()
        if path.suffix.lower() == ".db":
            with closing(connect(path, readonly=True)) as connection:
                return connection.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    except Exception:
        return None
    return None


def normalize_applications(payload: Any, job_map: dict[str, list[str]], job_ids: set[str], actions: dict[str, list[dict[str, Any]]], findings: dict[str, list[dict[str, Any]]], existing_ids: dict[str, set[str]]) -> list[dict[str, Any]]:
    records = []
    if isinstance(payload, dict):
        records = [{"jobId": key, **(value if isinstance(value, dict) else {})} for key, value in payload.items()]
    elif isinstance(payload, list):
        records = [value for value in payload if isinstance(value, dict)]
    normalized = []
    seen = set()
    for index, record in enumerate(records):
        source_id = str(record.get("jobId") or record.get("id") or "")
        resolved = job_map.get(source_id, [source_id])[0]
        if not resolved or resolved not in job_ids or resolved in seen:
            reason = "job relationship not found" if resolved not in job_ids else "duplicate application job ID"
            item = {"table": "applications", "jobId": source_id, "sourceIndex": index, "reason": reason}
            findings["invalidRecords" if resolved not in job_ids else "duplicates"].append(item)
            actions["applications"].append({"action": "skip", **item})
            continue
        seen.add(resolved)
        item = dict(record)
        item["jobId"] = resolved
        normalized.append(item)
        actions["applications"].append(action_for_id("applications", resolved, existing_ids, key="jobId"))
    return normalized


def normalize_candidates(payload: Any, companies_by_name: dict[str, list[str]], company_ids: set[str], actions: dict[str, list[dict[str, Any]]], findings: dict[str, list[dict[str, Any]]], existing_ids: dict[str, set[str]]) -> list[dict[str, Any]]:
    normalized = []
    used = set()
    for index, raw in enumerate(payload if isinstance(payload, list) else []):
        if not isinstance(raw, dict):
            item = {"table": "raw_job_candidates", "sourceIndex": index, "reason": "record is not an object"}
            findings["invalidRecords"].append(item)
            actions["raw_job_candidates"].append({"action": "skip", **item})
            continue
        company_id = str(raw.get("companyId") or "")
        if company_id not in company_ids:
            matches = companies_by_name.get(str(raw.get("companyName") or "").casefold(), [])
            company_id = matches[0] if len(matches) == 1 else ""
        basis = json.dumps(raw, sort_keys=True, ensure_ascii=True)
        candidate_id = "candidate-" + hashlib.sha256(f"{index}:{basis}".encode()).hexdigest()[:32]
        while candidate_id in used:
            candidate_id += "x"
        used.add(candidate_id)
        item = {"id": candidate_id, "companyId": company_id, "jobId": "", "companyName": str(raw.get("companyName") or ""), "candidateText": str(raw.get("candidateText") or ""), "candidateHref": str(raw.get("candidateHref") or ""), "rejectionReason": str(raw.get("rejectionReason") or ""), "payload": raw}
        normalized.append(item)
        actions["raw_job_candidates"].append(action_for_id("raw_job_candidates", candidate_id, existing_ids))
    return normalized


def normalize_utility_runs(paths: list[Path], actions: dict[str, list[dict[str, Any]]], existing_ids: dict[str, set[str]]) -> list[dict[str, Any]]:
    runs = []
    for path in paths:
        payload = read_json(path, [])
        for index, row in enumerate(payload if isinstance(payload, list) else []):
            basis = json.dumps(row, sort_keys=True, default=str)
            run_id = "utility-" + hashlib.sha256(f"{path.name}:{index}:{basis}".encode()).hexdigest()[:32]
            run = {"id": run_id, "utilityName": path.stem, "status": str(row.get("status") or row.get("auditStatus") or "") if isinstance(row, dict) else "", "payload": row}
            runs.append(run)
            actions["utility_runs"].append(action_for_id("utility_runs", run_id, existing_ids))
    return runs


def load_existing_ids(database_path: Path) -> dict[str, set[str]]:
    result = {"companies": set(), "jobs": set(), "applications": set(), "raw_job_candidates": set(), "utility_runs": set()}
    if not database_path.exists():
        return result
    try:
        with closing(connect(database_path, readonly=True)) as connection:
            for table, column in (("companies", "id"), ("jobs", "id"), ("applications", "job_id"), ("raw_job_candidates", "id"), ("utility_runs", "id")):
                result[table] = {row[0] for row in connection.execute(f"SELECT {column} FROM {table}")}
    except sqlite3.Error:
        pass
    return result


def action_for_id(table: str, record_id: str, existing_ids: dict[str, set[str]], *, key: str = "id") -> dict[str, Any]:
    return {"action": "update" if record_id in existing_ids.get(table, set()) else "create", "table": table, key: record_id}


def excel_company_to_api(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": str(row.get("Company ID") or ""), "name": str(row.get("Company Name") or ""),
        "industry": str(row.get("Industry") or "Financial Services"), "city": str(row.get("City") or ""),
        "state": str(row.get("State") or ""), "country": str(row.get("Country") or "United States"),
        "knownWebsite": str(row.get("Known Website") or ""), "officialWebsite": str(row.get("Official Website") or ""),
        "websiteDiscoveryMethod": str(row.get("Website Discovery Method") or ""),
        "websiteCandidateUrls": str(row.get("Website Candidate URLs") or ""),
        "websiteVerificationNotes": str(row.get("Website Verification Notes") or ""),
        "websiteVerified": bool_value(row.get("Website Verified")), "careersPageUrl": str(row.get("Careers Page URL") or ""),
        "jobBoardUrl": str(row.get("Job Board URL") or ""), "jobBoardDiscoveryMethod": str(row.get("Job Board Discovery Method") or "Not Found"),
        "jobsRssFeedUrl": str(row.get("Jobs RSS Feed URL") or ""), "jobPlatform": str(row.get("Job Platform") or ""),
        "feedFound": bool_value(row.get("Feed Found")), "searchStatus": str(row.get("Search Status") or "Needs Review"),
        "confidence": int_value(row.get("Confidence")), "lastChecked": str(row.get("Last Checked") or ""), "notes": str(row.get("Notes") or ""),
        "foundedYear": optional_int_value(row.get("Founded Year")),
        "totalAssets": optional_float_value(row.get("Total Assets")),
        "assetsAsOfDate": str(row.get("Assets As Of Date") or ""),
        "companyInfoLastChecked": str(row.get("Company Information Last Checked") or ""),
    }


def company_values(company: dict[str, Any], now: str) -> tuple[Any, ...]:
    return (company["id"],company["name"],company.get("industry","Financial Services"),company.get("city",""),company.get("state",""),company.get("country","United States"),company.get("knownWebsite",""),company.get("officialWebsite",""),company.get("websiteDiscoveryMethod",""),company.get("websiteCandidateUrls",""),company.get("websiteVerificationNotes",""),bool(company.get("websiteVerified")),company.get("careersPageUrl",""),company.get("jobBoardUrl",""),company.get("jobBoardDiscoveryMethod","Not Found"),company.get("jobsRssFeedUrl",""),company.get("jobPlatform",""),bool(company.get("feedFound")),company.get("searchStatus","Needs Review"),company.get("confidence",0),company.get("lastChecked",""),company.get("notes",""),company.get("foundedYear"),company.get("totalAssets"),company.get("assetsAsOfDate",""),company.get("companyInfoLastChecked",""),now,now)


def job_values(job: dict[str, Any], now: str) -> tuple[Any, ...]:
    return (job["id"],job.get("legacyId",job["id"]),job.get("companyId") or None,job.get("companyName",""),job.get("title",""),job.get("location",""),job.get("workType","Not Listed"),job.get("payMin"),job.get("payMax"),job.get("payText",""),job.get("payPeriod","unknown"),job.get("payCurrency","USD"),job.get("postedDate",""),job.get("sourceUrl",""),job.get("jobPlatform",""),job.get("description",""),job.get("descriptionSnippet",""),job.get("collectedAt",""),job.get("status","Open"),job.get("roleType","UNKNOWN"),job.get("roleTypeReason",""),json.dumps(job.get("rawData",{}),sort_keys=True),job.get("firstSeenAt") or now,now,now)


def application_values(application: dict[str, Any], now: str) -> tuple[Any, ...]:
    return (application["jobId"],bool(application.get("applied")),application.get("applicationStatus","Interested"),application.get("dateApplied",""),application.get("followUpDate",""),application.get("notes",""),bool(application.get("notInterested")),json.dumps(application,sort_keys=True),now)


def repaired_job_id(legacy_id: str, job: dict[str, Any], used: set[str]) -> str:
    basis = "|".join(str(job.get(key) or "") for key in ("companyId", "title", "sourceUrl", "location"))
    digest = hashlib.sha256(basis.encode()).hexdigest()[:12]
    candidate = f"{legacy_id}--{digest}"
    ordinal = 2
    while candidate in used:
        candidate = f"{legacy_id}--{digest}-{ordinal}"
        ordinal += 1
    return candidate


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return default


def bool_value(value: Any) -> bool:
    return value is True or str(value or "").strip().lower() in {"true", "yes", "1"}


def int_value(value: Any) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def optional_int_value(value: Any) -> int | None:
    try:
        return int(float(value)) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def optional_float_value(value: Any) -> float | None:
    try:
        return float(str(value).replace(",", "")) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def cleanup_sqlite_sidecars(database_path: Path) -> None:
    for suffix in ("-wal", "-shm", "-journal"):
        database_path.with_name(database_path.name + suffix).unlink(missing_ok=True)


def write_reports(report: dict[str, Any], json_path: Path, xlsx_path: Path) -> None:
    json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Migration Version", report["migrationVersion"]])
    summary.append(["Status", report["status"]])
    summary.append(["Backup Directory", report["backupDirectory"]])
    for group, values in report["counts"].items():
        if isinstance(values, dict):
            for key, value in values.items():
                summary.append([f"{group}.{key}", json.dumps(value) if isinstance(value, dict) else value])
    actions_sheet = workbook.create_sheet("Actions")
    actions_sheet.append(["Table", "Action", "ID", "Details"])
    for table, records in report["actions"].items():
        for record in records:
            record_id = record.get("id") or record.get("jobId") or record.get("originalId") or ""
            actions_sheet.append([table, record.get("action", ""), record_id, json.dumps(record, sort_keys=True)])
    findings_sheet = workbook.create_sheet("Findings")
    findings_sheet.append(["Category", "Table", "ID", "Details"])
    for category, records in report["findings"].items():
        for record in records:
            findings_sheet.append([category, record.get("table", ""), record.get("id") or record.get("originalId") or "", json.dumps(record, sort_keys=True)])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(xlsx_path)
