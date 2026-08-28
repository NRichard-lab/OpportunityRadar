from __future__ import annotations

import io
import json
import os
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from config import (
    MAX_IMPORT_ARCHIVE_FILES,
    MAX_IMPORT_ROWS,
    MAX_IMPORT_UNCOMPRESSED_BYTES,
    MAX_IMPORT_UPLOAD_BYTES,
    MAX_IMPORT_WORKSHEETS,
    OUTPUT_COLUMNS,
)


MAX_IMPORT_COLUMNS = len(OUTPUT_COLUMNS)
MAX_IMPORT_CELLS = MAX_IMPORT_COLUMNS * (MAX_IMPORT_ROWS + MAX_IMPORT_WORKSHEETS)
# Keep JSON validation independent of interpreter-specific decoder recursion
# behavior. Opportunity Radar imports have a shallow, known schema, so deeply
# nested documents are malformed for this interface even on Python builds whose
# C JSON decoder can parse beyond the process recursion limit.
MAX_IMPORT_JSON_NESTING = 64


def validate_import_upload(filename: str, contents: bytes) -> str:
    suffix = Path(filename).suffix.casefold()
    if suffix not in {".json", ".xlsx"}:
        raise ValueError("Import supports JSON and Excel (.xlsx) files.")
    if not contents:
        raise ValueError("The selected import file is empty.")
    if len(contents) > MAX_IMPORT_UPLOAD_BYTES:
        raise ValueError("The import exceeds the 25 MiB upload limit.")
    if suffix == ".json":
        _validate_json_bytes(contents)
    else:
        if not contents.startswith(b"PK\x03\x04"):
            raise ValueError("The file contents do not match the XLSX extension.")
        _validate_xlsx_archive(io.BytesIO(contents))
    return suffix


def stage_import_upload(directory: Path, filename: str, contents: bytes) -> Path:
    suffix = validate_import_upload(filename, contents)
    target_directory = Path(directory)
    target_directory.mkdir(parents=True, exist_ok=True)
    staged_path: Path | None = None
    try:
        descriptor, raw_path = tempfile.mkstemp(
            dir=target_directory,
            prefix=".opportunity-radar-import-",
            suffix=suffix,
        )
        staged_path = Path(raw_path)
        with os.fdopen(descriptor, "wb") as handle:
            handle.write(contents)
            handle.flush()
            os.fsync(handle.fileno())
        return staged_path
    except Exception:
        if staged_path is not None:
            staged_path.unlink(missing_ok=True)
        raise


def validate_staged_import(path: Path) -> None:
    source = Path(path)
    if source.stat().st_size > MAX_IMPORT_UPLOAD_BYTES:
        raise ValueError("The import exceeds the 25 MiB upload limit.")
    if source.suffix.casefold() == ".json":
        _validate_json_bytes(source.read_bytes())
        return
    if source.suffix.casefold() != ".xlsx":
        raise ValueError("Import supports JSON and Excel (.xlsx) files.")
    with source.open("rb") as handle:
        _validate_xlsx_archive(handle)
    try:
        workbook = load_workbook(source, read_only=True, data_only=False, keep_links=False)
        if len(workbook.worksheets) > MAX_IMPORT_WORKSHEETS:
            raise ValueError(f"Excel imports may contain at most {MAX_IMPORT_WORKSHEETS} worksheets.")
        _enforce_xlsx_shape_limits(workbook)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("The Excel import is malformed or could not be read.") from exc
    finally:
        if "workbook" in locals():
            workbook.close()


def enforce_record_limit(company_payloads: object, job_payloads: object) -> tuple[list[dict], list[dict]]:
    if not isinstance(company_payloads, list) or not isinstance(job_payloads, list):
        raise ValueError("Import record collections must be JSON arrays.")
    if len(company_payloads) + len(job_payloads) > MAX_IMPORT_ROWS:
        raise ValueError(f"Imports may contain at most {MAX_IMPORT_ROWS} records.")
    if any(not isinstance(item, dict) for item in [*company_payloads, *job_payloads]):
        raise ValueError("Every imported record must be a JSON object.")
    return company_payloads, job_payloads


def _validate_json_bytes(contents: bytes) -> None:
    try:
        document = contents.decode("utf-8-sig")
        _enforce_json_nesting_limit(document)
        payload = json.loads(document)
    except (UnicodeDecodeError, RecursionError, ValueError) as exc:
        raise ValueError("The JSON import is malformed or could not be read.") from exc
    if isinstance(payload, dict):
        enforce_record_limit(payload.get("companies", []), payload.get("jobs", []))
    elif isinstance(payload, list):
        if len(payload) > MAX_IMPORT_ROWS or any(not isinstance(item, dict) for item in payload):
            raise ValueError(f"Imports may contain at most {MAX_IMPORT_ROWS} object records.")
    else:
        raise ValueError("The selected JSON file does not contain company or job records.")


def _enforce_json_nesting_limit(document: str) -> None:
    """Reject excessive structural nesting without interpreting string data."""
    depth = 0
    in_string = False
    escaped = False
    for character in document:
        if in_string:
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == '"':
                in_string = False
            continue
        if character == '"':
            in_string = True
        elif character in "[{":
            depth += 1
            if depth > MAX_IMPORT_JSON_NESTING:
                raise RecursionError("JSON nesting exceeds the safe import limit.")
        elif character in "]}" and depth:
            depth -= 1


def _enforce_xlsx_shape_limits(workbook: object) -> None:
    total_rows = 0
    populated_cells = 0
    for sheet in workbook.worksheets:
        # Read-only openpyxl otherwise trusts the attacker-controlled worksheet
        # <dimension> value and may stop before later row elements are parsed.
        sheet.reset_dimensions()
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if len(row) > MAX_IMPORT_COLUMNS:
                raise ValueError(f"Excel imports may contain at most {MAX_IMPORT_COLUMNS} columns.")
            populated_cells += sum(value is not None for value in row)
            if populated_cells > MAX_IMPORT_CELLS:
                raise ValueError("The Excel import contains too many populated cells.")
            if row_number > 1:
                total_rows += 1
                if total_rows > MAX_IMPORT_ROWS:
                    raise ValueError(f"Imports may contain at most {MAX_IMPORT_ROWS} records.")


def _validate_xlsx_archive(source: object) -> None:
    try:
        with zipfile.ZipFile(source) as archive:
            entries = archive.infolist()
            names = {entry.filename for entry in entries}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ValueError("The XLSX file is missing required workbook content.")
            if len(entries) > MAX_IMPORT_ARCHIVE_FILES:
                raise ValueError("The XLSX archive contains too many files.")
            worksheets = [name for name in names if name.startswith("xl/worksheets/") and name.endswith(".xml")]
            if len(worksheets) > MAX_IMPORT_WORKSHEETS:
                raise ValueError(f"Excel imports may contain at most {MAX_IMPORT_WORKSHEETS} worksheets.")
            expanded_size = sum(max(0, entry.file_size) for entry in entries)
            if expanded_size > MAX_IMPORT_UNCOMPRESSED_BYTES:
                raise ValueError("The XLSX archive expands beyond the safe processing limit.")
            for entry in entries:
                if entry.file_size > 1024 * 1024 and entry.file_size > max(1, entry.compress_size) * 200:
                    raise ValueError("The XLSX archive has an unsafe compression ratio.")
            if archive.testzip() is not None:
                raise ValueError("The XLSX archive is corrupted.")
    except (zipfile.BadZipFile, EOFError, RuntimeError) as exc:
        raise ValueError("The XLSX file is malformed or could not be read.") from exc
