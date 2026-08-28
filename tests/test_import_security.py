from __future__ import annotations

import json
import re
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook

from backend.import_security import (
    MAX_IMPORT_COLUMNS,
    MAX_IMPORT_JSON_NESTING,
    _validate_xlsx_archive,
    stage_import_upload,
    validate_import_upload,
    validate_staged_import,
)


def make_xlsx(*, rows: int = 1, worksheets: int = 1) -> bytes:
    workbook = Workbook()
    workbook.active.append(["Company Name"])
    for row_number in range(rows):
        workbook.active.append([f"Company {row_number}"])
    for index in range(1, worksheets):
        sheet = workbook.create_sheet(f"Sheet {index + 1}")
        sheet.append(["Company Name"])
    contents = BytesIO()
    workbook.save(contents)
    workbook.close()
    return contents.getvalue()


def make_xlsx_with_far_cell() -> bytes:
    workbook = Workbook()
    workbook.active.append(["Company Name"])
    workbook.active.append(["Example"])
    workbook.active["XFD2"] = "unexpected far cell"
    contents = BytesIO()
    workbook.save(contents)
    workbook.close()
    return contents.getvalue()


def replace_first_worksheet_dimension(contents: bytes, dimension: str) -> bytes:
    source = BytesIO(contents)
    destination = BytesIO()
    replaced = False
    with ZipFile(source) as input_archive, ZipFile(destination, "w", ZIP_DEFLATED) as output_archive:
        for entry in input_archive.infolist():
            data = input_archive.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                data, substitutions = re.subn(
                    rb'<dimension ref="[^"]+"\s*/>',
                    f'<dimension ref="{dimension}"/>'.encode(),
                    data,
                    count=1,
                )
                replaced = substitutions == 1
            output_archive.writestr(entry, data)
    if not replaced:
        raise AssertionError("The synthetic workbook did not contain a worksheet dimension.")
    return destination.getvalue()


class _FakeArchive:
    def __init__(self, entries: list[SimpleNamespace]) -> None:
        self._entries = entries

    def __enter__(self) -> "_FakeArchive":
        return self

    def __exit__(self, *_args: object) -> None:
        return None

    def infolist(self) -> list[SimpleNamespace]:
        return self._entries

    def testzip(self) -> None:
        return None


class ImportUploadSecurityTests(unittest.TestCase):
    def test_accepts_valid_json_and_xlsx(self) -> None:
        payload = json.dumps({"companies": [{"name": "Example"}], "jobs": []}).encode()
        self.assertEqual(validate_import_upload("records.json", payload), ".json")
        self.assertEqual(validate_import_upload("records.XLSX", make_xlsx()), ".xlsx")

    def test_rejects_upload_over_configured_limit_without_large_allocation(self) -> None:
        with patch("backend.import_security.MAX_IMPORT_UPLOAD_BYTES", 8):
            with self.assertRaisesRegex(ValueError, "upload limit"):
                validate_import_upload("records.json", b"[{}, {}, {}]")

    def test_rejects_signature_mismatch_and_malformed_archive(self) -> None:
        with self.assertRaisesRegex(ValueError, "do not match the XLSX extension"):
            validate_import_upload("records.xlsx", b'{"companies": []}')
        with self.assertRaisesRegex(ValueError, "XLSX file is malformed"):
            validate_import_upload("records.xlsx", b"PK\x03\x04not-a-zip")

    def test_normalizes_xlsx_archive_parser_failures(self) -> None:
        for exception_type in (RuntimeError, NotImplementedError, EOFError):
            with self.subTest(exception_type=exception_type.__name__):
                secret_error = exception_type("sensitive archive parser detail")
                with patch("backend.import_security.zipfile.ZipFile", side_effect=secret_error):
                    with self.assertRaises(ValueError) as raised:
                        _validate_xlsx_archive(BytesIO(b"synthetic"))
                self.assertEqual(str(raised.exception), "The XLSX file is malformed or could not be read.")
                self.assertNotIn("sensitive", str(raised.exception))

    def test_rejects_json_record_limit(self) -> None:
        payload = json.dumps([{}, {}]).encode()
        with patch("backend.import_security.MAX_IMPORT_ROWS", 1):
            with self.assertRaisesRegex(ValueError, "at most 1"):
                validate_import_upload("records.json", payload)

    def test_deeply_nested_json_is_reported_as_safe_malformed_input(self) -> None:
        payload = (b"[" * (MAX_IMPORT_JSON_NESTING + 1)) + b"{}" + (
            b"]" * (MAX_IMPORT_JSON_NESTING + 1)
        )
        with self.assertRaises(ValueError) as raised:
            validate_import_upload("records.json", payload)
        self.assertEqual(str(raised.exception), "The JSON import is malformed or could not be read.")
        self.assertIsInstance(raised.exception.__cause__, RecursionError)

    def test_json_nesting_guard_ignores_brackets_inside_strings(self) -> None:
        bracket_text = ("[" * (MAX_IMPORT_JSON_NESTING + 1)) + (
            "]" * (MAX_IMPORT_JSON_NESTING + 1)
        )
        payload = json.dumps({"companies": [{"name": bracket_text}], "jobs": []}).encode()
        self.assertEqual(validate_import_upload("records.json", payload), ".json")

    def test_json_decoder_value_error_is_reported_as_safe_malformed_input(self) -> None:
        secret_error = "numeric parser detail from C:/private/import.json"
        with patch("backend.import_security.json.loads", side_effect=ValueError(secret_error)):
            with self.assertRaises(ValueError) as raised:
                validate_import_upload("records.json", b"[1]")
        self.assertEqual(str(raised.exception), "The JSON import is malformed or could not be read.")
        self.assertNotIn(secret_error, str(raised.exception))

    def test_rejects_archive_file_and_expanded_size_limits(self) -> None:
        contents = make_xlsx()
        with patch("backend.import_security.MAX_IMPORT_ARCHIVE_FILES", 1):
            with self.assertRaisesRegex(ValueError, "too many files"):
                validate_import_upload("records.xlsx", contents)
        with patch("backend.import_security.MAX_IMPORT_UNCOMPRESSED_BYTES", 1):
            with self.assertRaisesRegex(ValueError, "expands beyond"):
                validate_import_upload("records.xlsx", contents)

    def test_rejects_unsafe_archive_ratio_without_large_allocation(self) -> None:
        entries = [
            SimpleNamespace(filename="[Content_Types].xml", file_size=1, compress_size=1),
            SimpleNamespace(filename="xl/workbook.xml", file_size=1, compress_size=1),
            SimpleNamespace(
                filename="xl/worksheets/sheet1.xml",
                file_size=(1024 * 1024) + 1,
                compress_size=1,
            ),
        ]
        with patch("backend.import_security.zipfile.ZipFile", return_value=_FakeArchive(entries)):
            with self.assertRaisesRegex(ValueError, "unsafe compression ratio"):
                _validate_xlsx_archive(BytesIO(b"synthetic"))

    def test_rejects_worksheet_limit_during_archive_validation(self) -> None:
        with patch("backend.import_security.MAX_IMPORT_WORKSHEETS", 1):
            with self.assertRaisesRegex(ValueError, "at most 1 worksheets"):
                validate_import_upload("records.xlsx", make_xlsx(worksheets=2))

    def test_rejects_row_limit_when_validating_staged_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            staged = stage_import_upload(Path(temp_dir), "records.xlsx", make_xlsx(rows=2))
            try:
                with patch("backend.import_security.MAX_IMPORT_ROWS", 1):
                    with self.assertRaisesRegex(ValueError, "at most 1 records"):
                        validate_staged_import(staged)
            finally:
                staged.unlink(missing_ok=True)

    def test_accepts_valid_staged_workbook_with_supported_dimensions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            staged = stage_import_upload(Path(temp_dir), "records.xlsx", make_xlsx(rows=2))
            try:
                validate_staged_import(staged)
            finally:
                staged.unlink(missing_ok=True)

    def test_rejects_rows_hidden_by_forged_worksheet_dimension(self) -> None:
        contents = replace_first_worksheet_dimension(make_xlsx(rows=3), "A1:A1")
        with tempfile.TemporaryDirectory() as temp_dir:
            staged = stage_import_upload(Path(temp_dir), "records.xlsx", contents)
            try:
                with patch("backend.import_security.MAX_IMPORT_ROWS", 1):
                    with self.assertRaisesRegex(ValueError, "at most 1 records"):
                        validate_staged_import(staged)
            finally:
                staged.unlink(missing_ok=True)

    def test_rejects_far_cell_hidden_by_forged_worksheet_dimension(self) -> None:
        contents = replace_first_worksheet_dimension(make_xlsx_with_far_cell(), "A1:A2")
        with tempfile.TemporaryDirectory() as temp_dir:
            staged = stage_import_upload(Path(temp_dir), "records.xlsx", contents)
            try:
                with self.assertRaisesRegex(ValueError, f"at most {MAX_IMPORT_COLUMNS} columns"):
                    validate_staged_import(staged)
            finally:
                staged.unlink(missing_ok=True)

    def test_rejects_populated_cell_count_over_configured_bound(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            staged = stage_import_upload(Path(temp_dir), "records.xlsx", make_xlsx(rows=2))
            try:
                with patch("backend.import_security.MAX_IMPORT_CELLS", 2):
                    with self.assertRaisesRegex(ValueError, "too many populated cells"):
                        validate_staged_import(staged)
            finally:
                staged.unlink(missing_ok=True)

    def test_staging_uses_random_names_and_caller_cleanup_removes_them(self) -> None:
        payload = b'{"companies": [], "jobs": []}'
        with tempfile.TemporaryDirectory() as temp_dir:
            staging_root = Path(temp_dir)
            first = stage_import_upload(staging_root, "../../customer-records.json", payload)
            second = stage_import_upload(staging_root, "../../customer-records.json", payload)
            self.assertNotEqual(first, second)
            self.assertEqual(first.parent, staging_root)
            self.assertTrue(first.name.startswith(".opportunity-radar-import-"))
            self.assertNotIn("customer-records", first.name)
            self.assertEqual(first.read_bytes(), payload)
            first.unlink()
            second.unlink()
            self.assertEqual(list(staging_root.iterdir()), [])

    def test_staging_failure_removes_partial_random_file(self) -> None:
        payload = b'{"companies": [], "jobs": []}'
        with tempfile.TemporaryDirectory() as temp_dir:
            staging_root = Path(temp_dir)
            with patch("backend.import_security.os.fsync", side_effect=OSError("disk failure")):
                with self.assertRaisesRegex(OSError, "disk failure"):
                    stage_import_upload(staging_root, "records.json", payload)
            self.assertEqual(list(staging_root.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
