from __future__ import annotations

import tempfile
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from backend.file_security import atomic_save_workbook, sanitize_spreadsheet_value
from config import OUTPUT_COLUMNS
from excel_tools import is_safe_spreadsheet_hyperlink, read_company_rows, write_results
from job_tools import JOB_COLUMNS, JobRecord, write_jobs_xlsx


def make_workbook(marker: str) -> Workbook:
    workbook = Workbook()
    workbook.active["A1"] = marker
    return workbook


class _FailingWorkbook:
    def save(self, path: Path) -> None:
        Path(path).write_bytes(b"partial replacement")
        raise OSError("synthetic save failure")


class _ForgedRangeSheet:
    max_row = 1_048_576
    max_column = 16_384

    def __init__(self, *, fail_after_header: bool = False) -> None:
        self.fail_after_header = fail_after_header
        self.iteration_options: dict[str, object] = {}

    def iter_rows(self, **options):
        self.iteration_options = options
        yield ("Company Name", "Notes")
        if self.fail_after_header:
            raise RuntimeError("synthetic worksheet read failure")
        yield ("Example", "bounded")


class _ReadOnlyWorkbook:
    def __init__(self, sheet: _ForgedRangeSheet) -> None:
        self.active = sheet
        self.closed = False

    def close(self) -> None:
        self.closed = True


class SpreadsheetSecurityTests(unittest.TestCase):
    def test_company_reader_clamps_forged_dimensions_and_closes_workbook(self) -> None:
        sheet = _ForgedRangeSheet()
        workbook = _ReadOnlyWorkbook(sheet)
        source = Path("forged-range.xlsx")

        with patch("excel_tools.MAX_IMPORT_ROWS", 2), patch(
            "excel_tools.load_workbook", return_value=workbook
        ) as mocked_load:
            rows = read_company_rows(source)

        mocked_load.assert_called_once_with(source, data_only=True, read_only=True, keep_links=False)
        self.assertEqual(sheet.iteration_options["max_row"], 3)
        self.assertEqual(sheet.iteration_options["max_col"], len(OUTPUT_COLUMNS))
        self.assertEqual(rows[0]["Company Name"], "Example")
        self.assertTrue(workbook.closed)

    def test_company_reader_closes_workbook_when_iteration_fails(self) -> None:
        workbook = _ReadOnlyWorkbook(_ForgedRangeSheet(fail_after_header=True))
        with patch("excel_tools.load_workbook", return_value=workbook):
            with self.assertRaisesRegex(RuntimeError, "worksheet read failure"):
                read_company_rows(Path("broken.xlsx"))
        self.assertTrue(workbook.closed)

    def test_atomic_save_replaces_target_with_complete_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "snapshot.xlsx"
            target.write_bytes(b"old contents")
            atomic_save_workbook(target, make_workbook("complete"))

            workbook = load_workbook(target, read_only=True)
            try:
                self.assertEqual(workbook.active["A1"].value, "complete")
            finally:
                workbook.close()
            self.assertEqual(list(Path(temp_dir).glob(".snapshot.xlsx.*.xlsx.tmp")), [])

    def test_failed_save_retains_existing_target_and_cleans_temporary_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "snapshot.xlsx"
            original = b"known-good-original"
            target.write_bytes(original)

            with self.assertRaisesRegex(OSError, "synthetic save failure"):
                atomic_save_workbook(target, _FailingWorkbook())

            self.assertEqual(target.read_bytes(), original)
            self.assertEqual(list(Path(temp_dir).glob(".snapshot.xlsx.*.xlsx.tmp")), [])

    def test_failed_atomic_replace_retains_existing_target_and_cleans_temporary_file(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "snapshot.xlsx"
            original = b"known-good-original"
            target.write_bytes(original)

            with patch("backend.file_security.os.replace", side_effect=OSError("replace failure")):
                with self.assertRaisesRegex(OSError, "replace failure"):
                    atomic_save_workbook(target, make_workbook("new"))

            self.assertEqual(target.read_bytes(), original)
            self.assertEqual(list(Path(temp_dir).glob(".snapshot.xlsx.*.xlsx.tmp")), [])

    def test_concurrent_writers_leave_one_complete_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "snapshot.xlsx"
            markers = [f"writer-{index}" for index in range(8)]

            with ThreadPoolExecutor(max_workers=4) as executor:
                list(executor.map(lambda marker: atomic_save_workbook(target, make_workbook(marker)), markers))

            workbook = load_workbook(target, read_only=True)
            try:
                self.assertIn(workbook.active["A1"].value, markers)
            finally:
                workbook.close()
            self.assertEqual(list(Path(temp_dir).glob(".snapshot.xlsx.*.xlsx.tmp")), [])

    def test_formula_prefixes_and_control_characters_are_safe_in_company_export(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "companies.xlsx"
            write_results(
                target,
                [{"Company Name": "Example", "Notes": "\x00  =HYPERLINK(\"https://evil.test\")"}],
            )
            workbook = load_workbook(target, data_only=False)
            try:
                sheet = workbook.active
                notes_column = OUTPUT_COLUMNS.index("Notes") + 1
                cell = sheet.cell(row=2, column=notes_column)
                self.assertEqual(cell.value, "'  =HYPERLINK(\"https://evil.test\")")
                self.assertEqual(cell.data_type, "s")
            finally:
                workbook.close()

    def test_formula_prefixes_are_safe_in_job_export(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "jobs.xlsx"
            write_jobs_xlsx(
                target,
                [JobRecord(id="job-1", companyId="company-1", companyName="Example", title="+cmd|' /C calc'!A0")],
            )
            workbook = load_workbook(target, data_only=False)
            try:
                cell = workbook.active.cell(row=2, column=JOB_COLUMNS.index("title") + 1)
                self.assertEqual(cell.value, "'+cmd|' /C calc'!A0")
                self.assertEqual(cell.data_type, "s")
            finally:
                workbook.close()

    def test_only_safe_http_urls_become_hyperlinks(self) -> None:
        urls = [
            "https://safe.example/path",
            "javascript:alert(1)",
            "file:///etc/passwd",
            "https://user:secret@safe.example/private",
            "https://[",
        ]
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "companies.xlsx"
            write_results(
                target,
                [
                    {"Company Name": f"Company {index}", "Known Website": url}
                    for index, url in enumerate(urls)
                ],
            )
            workbook = load_workbook(target, data_only=False)
            try:
                sheet = workbook.active
                url_column = OUTPUT_COLUMNS.index("Known Website") + 1
                self.assertEqual(sheet.cell(row=2, column=url_column).hyperlink.target, urls[0])
                for row in range(3, len(urls) + 2):
                    self.assertIsNone(sheet.cell(row=row, column=url_column).hyperlink)
            finally:
                workbook.close()

    def test_malformed_url_parsing_is_fail_closed(self) -> None:
        self.assertEqual(sanitize_spreadsheet_value("https://["), "'https://[")
        self.assertFalse(is_safe_spreadsheet_hyperlink("https://["))


if __name__ == "__main__":
    unittest.main()
