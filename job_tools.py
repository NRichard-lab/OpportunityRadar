from __future__ import annotations

import json
import logging
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook

from backend.file_security import atomic_save_workbook, atomic_write_text, sanitize_spreadsheet_value
from config import (
    APP_ENABLE_BROWSER_JOBS,
    APP_MAX_BROWSER_WORKERS,
    APP_MAX_HTTP_WORKERS,
    APP_WRITE_FRONTEND_MIRRORS,
    DEFAULT_FRONTEND_JOBS_JSON,
    LOG_DIR,
    OUTPUT_DIR,
)
from excel_tools import read_company_rows, stable_company_id
from job_enrichment import classify_role_type, extract_pay_info
from job_validation import is_valid_job_title, rejection_reason
from job_platforms import detect_job_platform

logger = logging.getLogger(__name__)


JOB_COLUMNS = [
    "id",
    "companyId",
    "companyName",
    "title",
    "location",
    "workType",
    "payMin",
    "payMax",
    "payText",
    "payPeriod",
    "payCurrency",
    "postedDate",
    "sourceUrl",
    "jobPlatform",
    "description",
    "descriptionSnippet",
    "collectedAt",
    "status",
    "roleType",
    "roleTypeReason",
]

DIAGNOSTIC_COLUMNS = [
    "companyName",
    "jobPlatform",
    "jobBoardUrl",
    "finalUrlAfterRedirect",
    "sourceUrlUsed",
    "sourceTypeUsed",
    "collectorSelected",
    "collectorSelectionReason",
    "playwrightUsed",
    "candidateJobElementsFound",
    "candidateElementsRejected",
    "validJobsSaved",
    "status",
    "outcome",
    "dataDisposition",
    "errorMessage",
    "durationSeconds",
    "averageSecondsPerCompany",
    "estimatedSecondsRemaining",
    "notes",
    "payExtractionSource",
    "payCandidateText",
    "payPatternMatched",
    "parsedPayMin",
    "parsedPayMax",
    "parsedPayPeriod",
    "firstCandidateTitles",
    "firstCandidateUrls",
    "rejectionReasons",
]

REJECTED_CANDIDATE_COLUMNS = [
    "companyId",
    "companyName",
    "jobBoardUrl",
    "finalUrlAfterRedirect",
    "collectorUsed",
    "candidateText",
    "candidateHref",
    "rejectionReason",
    "surroundingTextSnippet",
    "detailPageAttempted",
    "detailPageTitleFound",
]

INVALID_JOB_BOARD_URL_PARTS = [
    "loan",
    "loan-application",
    "mortgage",
    "membership",
    "member-application",
    "account-opening",
    "online-banking",
    "login",
    "credit-card",
    "benefits",
    "culture",
    "locations",
    "privacy",
    "terms",
]


@dataclass
class JobRecord:
    id: str
    companyId: str
    companyName: str
    title: str
    location: str = ""
    workType: str = "Not Listed"
    payMin: int | None = None
    payMax: int | None = None
    payText: str = ""
    payPeriod: str = "unknown"
    payCurrency: str = "USD"
    postedDate: str = ""
    sourceUrl: str = ""
    jobPlatform: str = ""
    description: str = ""
    descriptionSnippet: str = ""
    collectedAt: str = ""
    status: str = "Open"
    roleType: str = "UNKNOWN"
    roleTypeReason: str = ""
    rawData: dict[str, Any] = field(default_factory=dict)


@dataclass
class CollectionDiagnostic:
    companyName: str
    officialWebsite: str = ""
    careersPageUrl: str = ""
    jobBoardUrl: str = ""
    jobPlatform: str = ""
    finalUrlAfterRedirect: str = ""
    sourceUrlUsed: str = ""
    sourceTypeUsed: str = ""
    collectorSelected: str = ""
    collectorSelectionReason: str = ""
    playwrightUsed: bool = False
    candidateJobElementsFound: int = 0
    candidateElementsRejected: int = 0
    validJobsSaved: int = 0
    status: str = ""
    outcome: str = ""
    dataDisposition: str = ""
    errorMessage: str = ""
    durationSeconds: float = 0
    averageSecondsPerCompany: float = 0
    estimatedSecondsRemaining: float = 0
    notes: str = ""
    payExtractionSource: str = ""
    payCandidateText: str = ""
    payPatternMatched: str = ""
    parsedPayMin: int | None = None
    parsedPayMax: int | None = None
    parsedPayPeriod: str = ""
    firstCandidateTitles: str = ""
    firstCandidateUrls: str = ""
    rejectionReasons: str = ""


def collect_jobs(
    master_path: Path,
    jobs_json_path: Path,
    jobs_xlsx_path: Path,
    limit_companies: int | None = None,
    company_filter: str = "",
    company_filters: list[str] | None = None,
    company_ids: set[str] | None = None,
    max_workers: int = APP_MAX_HTTP_WORKERS,
    browser_workers: int = APP_MAX_BROWSER_WORKERS,
    delay_seconds: float = 1.0,
    dry_run: bool = False,
    debug_job_collection: bool = False,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancellation_event: Any = None,
) -> dict[str, Any]:
    from collectors.base import pick_collector

    started = time.monotonic()
    companies = read_company_rows(master_path)
    if company_ids is not None:
        companies = [company for company in companies if str(company.get("Company ID") or "") in company_ids]
    filters = company_filters or ([company_filter] if company_filter else [])
    if filters:
        companies = [company for company in companies if company_matches_filters(company, filters)]

    reviewed = len(companies)
    candidates: list[dict[str, Any]] = []
    diagnostics: list[CollectionDiagnostic] = []
    for company in companies:
        source_url, source_type = first_source(company)
        invalid_reason = invalid_job_board_reason(source_url) if source_type == "Job Board URL" else ""
        if source_url and not invalid_reason:
            candidates.append(company)
            continue
        diagnostics.append(build_skip_diagnostic(company, source_url, source_type, invalid_reason))
    skipped = reviewed - len(candidates)
    missing_skipped = sum(1 for diagnostic in diagnostics if diagnostic.status == "Missing Job Board URL")
    invalid_skipped = sum(1 for diagnostic in diagnostics if diagnostic.status == "Invalid Job Board URL")
    if limit_companies:
        omitted_candidates = candidates[limit_companies:]
        candidates = candidates[:limit_companies]
        diagnostics.extend(
            build_scope_retained_diagnostic(company, "Not selected by the collection limit.")
            for company in omitted_candidates
        )

    logger.info("Total companies reviewed: %s", reviewed)
    logger.info("Companies skipped because no usable source URL: %s", skipped)
    logger.info("Companies attempted: %s", len(candidates))

    if dry_run:
        logger.info("Dry run enabled; no job JSON or Excel snapshot will be written.")
        summary = build_summary(reviewed, skipped, missing_skipped, invalid_skipped, len(candidates), 0, 0, started)
        log_summary(summary)
        return summary

    all_jobs: list[JobRecord] = []
    successful_companies: list[dict[str, Any]] = []
    rejected_candidates: list[dict[str, Any]] = []
    errors = 0
    http_companies = []
    browser_companies = []

    for company in candidates:
        log_company_input(company)
        debug_dir = LOG_DIR / "job_collection_debug" if debug_job_collection else None
        collector = pick_collector(company, delay_seconds=delay_seconds, debug=debug_job_collection, debug_dir=debug_dir)
        if collector.requires_browser:
            browser_companies.append((company, collector))
        else:
            http_companies.append((company, collector))

    total_attempted = len(http_companies) + len(browser_companies)
    completed_counter = {"count": 0}

    http_workers = min(APP_MAX_HTTP_WORKERS, max(1, int(max_workers or 1)))
    browser_workers = min(APP_MAX_BROWSER_WORKERS, max(1, int(browser_workers or 1))) if APP_ENABLE_BROWSER_JOBS else 0

    for company, jobs, had_error, diagnostic, rejected in run_collection_group(http_companies, http_workers, total_attempted, completed_counter, started, "HTTP", progress_callback, cancellation_event):
        all_jobs.extend(jobs)
        errors += int(had_error)
        diagnostics.append(diagnostic)
        rejected_candidates.extend(rejected)
        if not had_error:
            successful_companies.append(company)

    if browser_workers:
        for company, jobs, had_error, diagnostic, rejected in run_collection_group(browser_companies, browser_workers, total_attempted, completed_counter, started, "Browser", progress_callback, cancellation_event):
            all_jobs.extend(jobs)
            errors += int(had_error)
            diagnostics.append(diagnostic)
            rejected_candidates.extend(rejected)
            if not had_error:
                successful_companies.append(company)
    else:
        for company, collector in browser_companies:
            diagnostics.append(build_retained_diagnostic(company, collector, "Browser collection is disabled."))

    if cancellation_event is not None and cancellation_event.is_set():
        raise InterruptedError("Job refresh cancelled.")
    all_jobs = [enrich_job_record(job) for job in all_jobs]
    final_jobs = merge_with_existing_jobs(jobs_json_path, all_jobs, successful_companies)
    write_jobs_json(jobs_json_path, final_jobs)
    write_jobs_xlsx(jobs_xlsx_path, final_jobs)
    write_diagnostics(diagnostics)
    write_rejected_candidates(rejected_candidates)
    summary = build_summary(
        reviewed, skipped, missing_skipped, invalid_skipped, len(candidates), len(all_jobs), errors, started,
        diagnostics=diagnostics,
    )
    log_summary(summary)
    return summary


def log_company_input(company: dict[str, Any]) -> None:
    logger.info("Company Name: %s", company.get("Company Name") or "")
    logger.info("Official Website: %s", company.get("Official Website") or "")
    logger.info("Careers Page URL: %s", company.get("Careers Page URL") or "")
    logger.info("Job Board URL: %s", company.get("Job Board URL") or "")
    logger.info("Job Platform: %s", company.get("Job Platform") or "")
    logger.info("Job Board Discovery Method: %s", company.get("Job Board Discovery Method") or "")
    logger.info("Search Status: %s", company.get("Search Status") or "")
    logger.info("Confidence: %s", company.get("Confidence") or "")
    logger.info("Last Checked: %s", company.get("Last Checked") or "")


def run_collection_group(
    collection_jobs: list[tuple[dict[str, Any], Any]],
    workers: int,
    total_attempted: int,
    completed_counter: dict[str, int],
    overall_started: float,
    worker_type: str,
    progress_callback: Callable[[int, int, str], None] | None = None,
    cancellation_event: Any = None,
):
    if not collection_jobs:
        return []
    results = []
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        futures = {
            executor.submit(collector.collect, company): (company, collector, time.monotonic())
            for company, collector in collection_jobs
        }
        for future in as_completed(futures):
            company, collector, started = futures[future]
            if cancellation_event is not None and cancellation_event.is_set():
                for pending in futures:
                    pending.cancel()
                raise InterruptedError("Job refresh cancelled.")
            try:
                jobs = [enrich_job_record(job) for job in future.result()]
                jobs = [job for job in jobs if is_valid_job_record(job)]
                candidates_seen = max(collector.candidate_count, collector.rejected_count + len(jobs))
                diagnostic = build_diagnostic(company, collector, candidates_seen, len(jobs), "", started)
                update_progress_diagnostic(diagnostic, total_attempted, completed_counter, overall_started)
                log_collection_progress(diagnostic, worker_type, total_attempted, completed_counter["count"])
                log_diagnostic(diagnostic)
                results.append((company, jobs, False, diagnostic, list(getattr(collector, "rejected_candidates", []))))
            except Exception as exc:
                logger.exception("Collector failed.")
                outcome = "timed-out" if isinstance(exc, TimeoutError) or "timeout" in type(exc).__name__.casefold() else "failed"
                diagnostic = build_diagnostic(
                    company, collector, collector.candidate_count, 0, safe_collection_error(exc), started,
                    outcome=outcome,
                )
                update_progress_diagnostic(diagnostic, total_attempted, completed_counter, overall_started)
                log_collection_progress(diagnostic, worker_type, total_attempted, completed_counter["count"])
                log_diagnostic(diagnostic)
                results.append((company, [], True, diagnostic, list(getattr(collector, "rejected_candidates", []))))
            if progress_callback:
                progress_callback(completed_counter["count"], total_attempted, str(company.get("Company Name") or ""))
    return results


def build_diagnostic(
    company: dict[str, Any],
    collector: Any,
    candidates_seen: int,
    jobs_saved: int,
    error_message: str,
    started: float,
    *,
    outcome: str | None = None,
) -> CollectionDiagnostic:
    source_url, source_type = collector.source_url(company)
    return CollectionDiagnostic(
        companyName=str(company.get("Company Name") or ""),
        officialWebsite=str(company.get("Official Website") or ""),
        careersPageUrl=str(company.get("Careers Page URL") or ""),
        jobBoardUrl=str(company.get("Job Board URL") or ""),
        jobPlatform=str(company.get("Job Platform") or ""),
        finalUrlAfterRedirect=str(getattr(collector, "final_url_after_redirect", "")),
        sourceUrlUsed=source_url,
        sourceTypeUsed=source_type,
        collectorSelected=collector.__class__.__name__,
        collectorSelectionReason=str(getattr(collector, "selection_reason", "")),
        playwrightUsed=bool(getattr(collector, "requires_browser", False)),
        candidateJobElementsFound=candidates_seen,
        candidateElementsRejected=int(getattr(collector, "rejected_count", 0)),
        validJobsSaved=jobs_saved,
        status=diagnostic_status(jobs_saved, error_message, collector),
        outcome=outcome or ("success" if jobs_saved else "success-empty"),
        dataDisposition="retained" if error_message else "replaced" if jobs_saved else "replaced-empty",
        errorMessage=error_message,
        durationSeconds=round(time.monotonic() - started, 2),
        notes=diagnostic_notes(collector, jobs_saved, error_message),
        payExtractionSource=str(getattr(collector, "last_pay_extraction", {}).get("source", "")) if hasattr(collector, "last_pay_extraction") else "",
        payCandidateText=str(getattr(collector, "last_pay_extraction", {}).get("candidateText", "")) if hasattr(collector, "last_pay_extraction") else "",
        payPatternMatched=str(getattr(collector, "last_pay_extraction", {}).get("pattern", "")) if hasattr(collector, "last_pay_extraction") else "",
        parsedPayMin=getattr(collector, "last_pay_extraction", {}).get("payMin") if hasattr(collector, "last_pay_extraction") else None,
        parsedPayMax=getattr(collector, "last_pay_extraction", {}).get("payMax") if hasattr(collector, "last_pay_extraction") else None,
        parsedPayPeriod=str(getattr(collector, "last_pay_extraction", {}).get("payPeriod", "")) if hasattr(collector, "last_pay_extraction") else "",
        firstCandidateTitles=" | ".join(sample.get("text", "") for sample in getattr(collector, "candidate_samples", [])),
        firstCandidateUrls=" | ".join(sample.get("url", "") for sample in getattr(collector, "candidate_samples", [])),
        rejectionReasons=" | ".join(
            f"{sample.get('text', '')}: {sample.get('reason', '')}"
            for sample in getattr(collector, "rejection_samples", [])
        ),
    )


def log_diagnostic(diagnostic: CollectionDiagnostic) -> None:
    logger.info("Company: %s", diagnostic.companyName)
    logger.info("Job Board URL: %s", diagnostic.jobBoardUrl)
    logger.info("Final URL after redirect: %s", diagnostic.finalUrlAfterRedirect)
    logger.info("Job Platform: %s", diagnostic.jobPlatform)
    logger.info("Source URL used: %s", diagnostic.sourceUrlUsed)
    logger.info("Source type used: %s", diagnostic.sourceTypeUsed)
    logger.info("Collector selected: %s", diagnostic.collectorSelected)
    logger.info("Collector selection reason: %s", diagnostic.collectorSelectionReason)
    logger.info("Playwright used: %s", diagnostic.playwrightUsed)
    logger.info("Candidate job elements found: %s", diagnostic.candidateJobElementsFound)
    logger.info("Candidate elements rejected: %s", diagnostic.candidateElementsRejected)
    logger.info("Valid jobs saved: %s", diagnostic.validJobsSaved)
    logger.info("Status: %s", diagnostic.status)
    if diagnostic.errorMessage:
        logger.info("Error message: %s", diagnostic.errorMessage)
    logger.info("Duration seconds: %s", diagnostic.durationSeconds)
    logger.info("Average seconds per company: %s", diagnostic.averageSecondsPerCompany)
    logger.info("Estimated time remaining: %ss", diagnostic.estimatedSecondsRemaining)
    if diagnostic.notes:
        logger.info("Notes: %s", diagnostic.notes)


def update_progress_diagnostic(
    diagnostic: CollectionDiagnostic,
    total_attempted: int,
    completed_counter: dict[str, int],
    overall_started: float,
) -> None:
    completed_counter["count"] += 1
    completed = completed_counter["count"]
    elapsed = max(0.001, time.monotonic() - overall_started)
    average = elapsed / completed
    diagnostic.averageSecondsPerCompany = round(average, 2)
    diagnostic.estimatedSecondsRemaining = round(max(0.0, average * (total_attempted - completed)), 2)


def log_collection_progress(
    diagnostic: CollectionDiagnostic,
    worker_type: str,
    total_attempted: int,
    completed: int,
) -> None:
    logger.info(
        "Processing %s of %s companies | Company: %s | Collector: %s | Worker type: %s | Jobs saved: %s | Duration seconds: %s | Average seconds per company: %s | Estimated time remaining: %ss",
        completed,
        total_attempted,
        diagnostic.companyName,
        diagnostic.collectorSelected,
        worker_type,
        diagnostic.validJobsSaved,
        diagnostic.durationSeconds,
        diagnostic.averageSecondsPerCompany,
        diagnostic.estimatedSecondsRemaining,
    )


def first_source(company: dict[str, Any]) -> tuple[str, str]:
    job_board_url = str(company.get("Job Board URL") or "").strip()
    if job_board_url:
        return job_board_url, "Job Board URL"
    rss_url = str(company.get("Jobs RSS Feed URL") or "").strip()
    feed_found = str(company.get("Feed Found") or "").strip().lower() in {"true", "yes", "1"}
    if rss_url and feed_found:
        return rss_url, "RSS Feed"
    return "", "None"


def company_matches_filters(company: dict[str, Any], filters: list[str]) -> bool:
    company_name = str(company.get("Company Name", "")).lower()
    return any(filter_value.strip().lower() in company_name for filter_value in filters if filter_value.strip())


def diagnostic_status(jobs_saved: int, error_message: str, collector: Any) -> str:
    if error_message:
        return "Collector Failed"
    if jobs_saved:
        return "Jobs Collected"
    if collector.__class__.__name__ == "GenericCollector" and not detect_job_platform(getattr(collector, "final_url_after_redirect", "")):
        return "No Jobs Found"
    return "No Jobs Found"


def diagnostic_notes(collector: Any, jobs_saved: int, error_message: str) -> str:
    if error_message:
        return f"Collector failed: {error_message}"
    if jobs_saved:
        return "Real job records were saved after validation."
    if "GenericCollector" in collector.__class__.__name__:
        return "No validated jobs saved from generic page parsing; review source page or add a platform-specific collector."
    return "No validated jobs saved; platform collector needs review."


def is_valid_job_record(job: JobRecord) -> bool:
    if not is_valid_job_title(job.title):
        logger.debug("Rejected job title %r: %s", job.title, rejection_reason(job.title))
        return False
    if not job.sourceUrl:
        return False
    if invalid_job_board_reason(job.sourceUrl):
        return False
    return bool(job.description or job.location or job.postedDate or job.sourceUrl or job.payText)


def enrich_job_record(job: JobRecord) -> JobRecord:
    if job.payText:
        pay_info = extract_pay_info(job.payText)
        if pay_info.get("payText"):
            job.payMin = pay_info.get("payMin")
            job.payMax = pay_info.get("payMax")
            job.payPeriod = str(pay_info.get("payPeriod") or "unknown")
            job.payCurrency = str(pay_info.get("payCurrency") or "USD")
    elif job.payMin is None and job.payMax is None:
        pay_info = extract_pay_info(" ".join([job.description, job.descriptionSnippet, job.rawData.get("rowText", "") if isinstance(job.rawData, dict) else ""]))
        if pay_info.get("payText"):
            job.payMin = pay_info.get("payMin")
            job.payMax = pay_info.get("payMax")
            job.payText = str(pay_info.get("payText") or "")
            job.payPeriod = str(pay_info.get("payPeriod") or "unknown")
            job.payCurrency = str(pay_info.get("payCurrency") or "USD")
            if isinstance(job.rawData, dict):
                job.rawData["payExtraction"] = pay_info
    if not job.roleType or job.roleType == "UNKNOWN":
        role = classify_role_type(job.title, job.description)
        job.roleType = role["roleType"]
        job.roleTypeReason = role["roleTypeReason"]
    return job


def merge_with_existing_jobs(path: Path, new_jobs: list[JobRecord], target_companies: list[dict[str, Any]]) -> list[JobRecord]:
    existing = load_existing_jobs(path)
    target_ids = {str(company.get("Company ID") or stable_company_id(company)) for company in target_companies}
    target_names = {str(company.get("Company Name") or "").lower() for company in target_companies}
    kept = [
        job for job in existing
        if (
            str(job.companyId or "") not in target_ids
            if str(job.companyId or "").strip()
            else str(job.companyName).lower() not in target_names
        )
    ]
    merged = [*kept, *new_jobs]
    unique: dict[str, JobRecord] = {}
    for job in merged:
        unique[dedupe_job_key(job)] = job
    return [job for job in unique.values() if is_valid_job_record(job)]


def load_existing_jobs(path: Path) -> list[JobRecord]:
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        logger.warning("Could not read existing jobs JSON for merge: %s", path)
        return []
    jobs: list[JobRecord] = []
    for item in payload if isinstance(payload, list) else []:
        if not isinstance(item, dict):
            continue
        fields = {field_name: item.get(field_name) for field_name in JobRecord.__dataclass_fields__}
        fields["rawData"] = fields.get("rawData") if isinstance(fields.get("rawData"), dict) else {}
        jobs.append(JobRecord(**fields))
    return jobs


def dedupe_job_key(job: JobRecord) -> str:
    if job.sourceUrl:
        return f"url:{job.sourceUrl.strip().lower()}"
    return f"fallback:{job.companyId}:{slug(job.title)}:{slug(job.location)}"


def slug(value: str) -> str:
    safe = "".join(char.lower() if char.isalnum() else "-" for char in str(value or "")).strip("-")
    while "--" in safe:
        safe = safe.replace("--", "-")
    return safe or "unknown"


def write_jobs_json(path: Path, jobs: list[JobRecord]) -> None:
    payload = [asdict(job) for job in jobs]
    atomic_write_text(path, json.dumps(payload, indent=2))
    mirror_jobs_json(path)
    logger.info("Wrote %s jobs to %s", len(jobs), path)


def mirror_jobs_json(path: Path) -> None:
    if not APP_WRITE_FRONTEND_MIRRORS:
        return
    if path.resolve() == DEFAULT_FRONTEND_JOBS_JSON.resolve():
        return
    if path.name != "jobs.json":
        return
    atomic_write_text(DEFAULT_FRONTEND_JOBS_JSON, path.read_text(encoding="utf-8"))
    logger.info("Mirrored jobs JSON to %s", DEFAULT_FRONTEND_JOBS_JSON)


def write_jobs_xlsx(path: Path, jobs: list[JobRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Jobs Snapshot"
    sheet.append(JOB_COLUMNS)
    for job in jobs:
        row = asdict(job)
        sheet.append([sanitize_spreadsheet_value(row.get(column, "")) for column in JOB_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    atomic_save_workbook(path, workbook)
    logger.info("Wrote jobs snapshot to %s", path)


def write_diagnostics(diagnostics: list[CollectionDiagnostic]) -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    json_path = LOG_DIR / "job_collection_diagnostics.json"
    xlsx_path = OUTPUT_DIR / "job_collection_diagnostics.xlsx"
    payload = [asdict(diagnostic) for diagnostic in diagnostics]
    atomic_write_text(json_path, json.dumps(payload, indent=2))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Job Collection Diagnostics"
    sheet.append(DIAGNOSTIC_COLUMNS)
    for row in payload:
        sheet.append([sanitize_spreadsheet_value(row.get(column, "")) for column in DIAGNOSTIC_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    atomic_save_workbook(xlsx_path, workbook)
    logger.info("Wrote job collection diagnostics to %s and %s", json_path, xlsx_path)


def write_rejected_candidates(
    rejected_candidates: list[dict[str, Any]],
    json_path: Path | None = None,
    xlsx_path: Path | None = None,
) -> None:
    json_path = json_path or LOG_DIR / "rejected_job_candidates.json"
    xlsx_path = xlsx_path or OUTPUT_DIR / "rejected_job_candidates.xlsx"
    json_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    atomic_write_text(json_path, json.dumps(rejected_candidates, indent=2))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rejected Job Candidates"
    sheet.append(REJECTED_CANDIDATE_COLUMNS)
    for row in rejected_candidates:
        sheet.append([sanitize_spreadsheet_value(row.get(column, "")) for column in REJECTED_CANDIDATE_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    atomic_save_workbook(xlsx_path, workbook)
    logger.info("Wrote rejected job candidates to %s and %s", json_path, xlsx_path)


def build_skip_diagnostic(
    company: dict[str, Any],
    source_url: str,
    source_type: str,
    invalid_reason: str,
) -> CollectionDiagnostic:
    missing = not source_url
    status = "Missing Job Board URL" if missing else "Invalid Job Board URL"
    notes = (
        "Skipped job collection because Job Board URL is missing."
        if missing
        else f"Skipped job collection because Job Board URL is invalid: {invalid_reason}"
    )
    return CollectionDiagnostic(
        companyName=str(company.get("Company Name") or ""),
        officialWebsite=str(company.get("Official Website") or ""),
        careersPageUrl=str(company.get("Careers Page URL") or ""),
        jobBoardUrl=str(company.get("Job Board URL") or ""),
        jobPlatform=str(company.get("Job Platform") or ""),
        sourceUrlUsed=source_url,
        sourceTypeUsed=source_type,
        collectorSelected="None",
        collectorSelectionReason=notes,
        status=status,
        outcome="stale",
        dataDisposition="retained",
        notes=notes,
    )


def invalid_job_board_reason(url: str) -> str:
    lowered = str(url or "").strip().lower()
    if not lowered:
        return ""
    if not lowered.startswith(("http://", "https://")):
        return "URL is not public HTTP(S)"
    for part in INVALID_JOB_BOARD_URL_PARTS:
        if part in lowered:
            return f"URL contains disallowed non-job pattern: {part}"
    if lowered.endswith((".pdf", ".doc", ".docx")):
        return "URL points to a document"
    return ""


def make_job_id(company: dict[str, Any], title: str, source_url: str) -> str:
    company_id = str(company.get("Company ID") or stable_company_id(company))
    basis = source_url or f"{company_id}-{title}"
    safe = "".join(char.lower() if char.isalnum() else "-" for char in basis).strip("-")
    while "--" in safe:
        safe = safe.replace("--", "-")
    return f"job-{safe[:140]}"


def build_summary(
    reviewed: int,
    skipped: int,
    missing_skipped: int,
    invalid_skipped: int,
    attempted: int,
    jobs_found: int,
    errors: int,
    started: float,
    *,
    diagnostics: list[CollectionDiagnostic] | None = None,
) -> dict[str, Any]:
    duration = max(0.001, time.monotonic() - started)
    results = diagnostics or []
    outcome_counts: dict[str, int] = {}
    for result in results:
        outcome_counts[result.outcome or "unknown"] = outcome_counts.get(result.outcome or "unknown", 0) + 1
    return {
        "total_companies_reviewed": reviewed,
        "companies_skipped_no_job_board_url": skipped,
        "companies_skipped_missing_job_board_url": missing_skipped,
        "companies_skipped_invalid_job_board_url": invalid_skipped,
        "companies_attempted": attempted,
        "jobs_found": jobs_found,
        "jobs_saved": jobs_found,
        "errors": errors,
        "outcome_counts": outcome_counts,
        "collection_results": [
            {
                "companyName": result.companyName,
                "outcome": result.outcome,
                "dataDisposition": result.dataDisposition,
                "status": result.status,
            }
            for result in results
        ],
        "duration_seconds": round(duration, 2),
        "average_seconds_per_company": round(duration / attempted, 2) if attempted else 0,
    }


def build_retained_diagnostic(company: dict[str, Any], collector: Any, reason: str) -> CollectionDiagnostic:
    source_url, source_type = collector.source_url(company)
    return CollectionDiagnostic(
        companyName=str(company.get("Company Name") or ""),
        jobBoardUrl=str(company.get("Job Board URL") or ""),
        jobPlatform=str(company.get("Job Platform") or ""),
        sourceUrlUsed=source_url,
        sourceTypeUsed=source_type,
        collectorSelected=collector.__class__.__name__,
        collectorSelectionReason=reason,
        playwrightUsed=True,
        status="Collection Disabled",
        outcome="stale",
        dataDisposition="retained",
        notes=reason,
    )


def build_scope_retained_diagnostic(company: dict[str, Any], reason: str) -> CollectionDiagnostic:
    source_url, source_type = first_source(company)
    return CollectionDiagnostic(
        companyName=str(company.get("Company Name") or ""),
        jobBoardUrl=str(company.get("Job Board URL") or ""),
        jobPlatform=str(company.get("Job Platform") or ""),
        sourceUrlUsed=source_url,
        sourceTypeUsed=source_type,
        collectorSelected="None",
        collectorSelectionReason=reason,
        status="Not Selected",
        outcome="stale",
        dataDisposition="retained",
        notes=reason,
    )


def safe_collection_error(exc: Exception) -> str:
    if isinstance(exc, TimeoutError) or "timeout" in type(exc).__name__.casefold():
        return "The collector timed out."
    return "The collector could not complete."


def log_summary(summary: dict[str, Any]) -> None:
    logger.info("Total companies reviewed: %s", summary["total_companies_reviewed"])
    logger.info("Companies skipped: Missing Job Board URL: %s", summary["companies_skipped_missing_job_board_url"])
    logger.info("Companies skipped: Invalid Job Board URL: %s", summary["companies_skipped_invalid_job_board_url"])
    logger.info("Companies skipped total: %s", summary["companies_skipped_no_job_board_url"])
    logger.info("Companies attempted: %s", summary["companies_attempted"])
    logger.info("Jobs found: %s", summary["jobs_found"])
    logger.info("Jobs saved: %s", summary["jobs_saved"])
    logger.info("Errors: %s", summary["errors"])
    logger.info("Duration: %ss", summary["duration_seconds"])
    logger.info("Average seconds per company: %s", summary["average_seconds_per_company"])
