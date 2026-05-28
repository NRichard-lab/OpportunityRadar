from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import DEFAULT_FRONTEND_COMPANIES_JSON, OUTPUT_COLUMNS


INPUT_COLUMNS = ["Company Name", "City", "State", "Known Website", "Notes"]
URL_COLUMNS = {
    "Known Website",
    "Official Website",
    "Website Discovery Method",
    "Website Candidate URLs",
    "Website Verification Notes",
    "Website Verified",
    "Careers Page URL",
    "Job Board URL",
    "Jobs RSS Feed URL",
}
COLUMN_MAP = {
    "Company ID": "id",
    "Company Name": "name",
    "City": "city",
    "State": "state",
    "Known Website": "knownWebsite",
    "Official Website": "officialWebsite",
    "Website Discovery Method": "websiteDiscoveryMethod",
    "Website Candidate URLs": "websiteCandidateUrls",
    "Website Verification Notes": "websiteVerificationNotes",
    "Website Verified": "websiteVerified",
    "Careers Page URL": "careersPageUrl",
    "Job Board URL": "jobBoardUrl",
    "Job Board Discovery Method": "jobBoardDiscoveryMethod",
    "Jobs RSS Feed URL": "jobsRssFeedUrl",
    "Job Platform": "jobPlatform",
    "Feed Found": "feedFound",
    "Search Status": "searchStatus",
    "Confidence": "confidence",
    "Last Checked": "lastChecked",
    "Notes": "notes",
}

logger = logging.getLogger(__name__)


def read_companies(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_indexes = {header: index for index, header in enumerate(headers)}
    if "Company Name" not in header_indexes:
        raise ValueError("Input workbook must include a 'Company Name' column.")

    companies: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {column: "" for column in INPUT_COLUMNS}
        for column in INPUT_COLUMNS:
            if column in header_indexes:
                value = row[header_indexes[column]]
                record[column] = str(value).strip() if value is not None else ""
        if record["Company Name"]:
            companies.append(record)
    return companies


def read_company_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {header: row[index] if index < len(row) else "" for index, header in enumerate(headers) if header}
        if record.get("Company Name"):
            records.append(normalize_company_row(record))
    return records


def write_results(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Enriched Companies"
    sheet.append(OUTPUT_COLUMNS)

    for result in rows:
        sheet.append([result.get(column, "") for column in OUTPUT_COLUMNS])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            header = sheet.cell(row=1, column=cell.column).value
            if header in URL_COLUMNS and cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 70)

    workbook.save(path)


def write_master(path: Path, rows: list[dict[str, Any]]) -> None:
    merged, stats = dedupe_company_rows(rows)
    write_results(path, merged)
    logger.info(
        "Master write complete: %s input rows, %s duplicate rows merged, %s master rows.",
        stats["input_rows"],
        stats["duplicate_rows"],
        stats["output_rows"],
    )


def update_master(master_path: Path, incoming_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    existing_rows = read_company_rows(master_path) if master_path.exists() else []
    merged_rows, stats = dedupe_company_rows([*existing_rows, *incoming_rows])
    write_results(master_path, merged_rows)
    logger.info(
        "Master merge complete: %s existing rows, %s incoming rows, %s duplicate rows merged, %s master rows.",
        len(existing_rows),
        len(incoming_rows),
        stats["duplicate_rows"],
        stats["output_rows"],
    )
    return merged_rows


def export_excel_to_json(input_path: Path, output_path: Path) -> int:
    rows = read_company_rows(input_path)
    merged_rows, stats = dedupe_company_rows(rows)
    logger.info(
        "Export dedupe complete: %s input rows, %s duplicate rows merged, %s companies exported.",
        stats["input_rows"],
        stats["duplicate_rows"],
        stats["output_rows"],
    )

    companies = []
    for row_data in merged_rows:
        company: dict[str, Any] = {}
        for excel_column, json_key in COLUMN_MAP.items():
            value = row_data.get(excel_column)
            if json_key == "feedFound":
                value = bool(value)
            elif json_key == "websiteVerified":
                value = str(value).strip().lower() in {"true", "yes", "1"}
            elif json_key == "confidence":
                value = normalize_confidence(value)
            elif json_key == "jobBoardDiscoveryMethod" and value is None:
                value = "Not Found"
            elif value is None:
                value = ""
            else:
                value = str(value)
            company[json_key] = value
        if not company.get("id"):
            company["id"] = stable_company_id(row_data)
        companies.append(company)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(companies, indent=2), encoding="utf-8")
    mirror_companies_json(output_path)
    return len(companies)


def mirror_companies_json(output_path: Path) -> None:
    if output_path.resolve() == DEFAULT_FRONTEND_COMPANIES_JSON.resolve():
        return
    if output_path.name != "companies.json":
        return
    DEFAULT_FRONTEND_COMPANIES_JSON.parent.mkdir(parents=True, exist_ok=True)
    DEFAULT_FRONTEND_COMPANIES_JSON.write_text(output_path.read_text(encoding="utf-8"), encoding="utf-8")
    logger.info("Mirrored companies JSON to %s", DEFAULT_FRONTEND_COMPANIES_JSON)


def normalize_confidence(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().lower()
    confidence_labels = {
        "high": 90,
        "medium": 65,
        "low": 35,
    }
    if text in confidence_labels:
        return confidence_labels[text]
    try:
        return int(float(text))
    except ValueError:
        return 0


def normalize_company_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {column: "" for column in OUTPUT_COLUMNS}
    for key, value in row.items():
        if key in normalized:
            normalized[key] = value if value is not None else ""
    if not normalized["Company Name"]:
        normalized["Company Name"] = str(row.get("name", "") or "")
    if not normalized["Company ID"]:
        normalized["Company ID"] = stable_company_id(normalized)
    if not normalized["Job Board Discovery Method"]:
        normalized["Job Board Discovery Method"] = "Not Found"
    if not normalized["Website Discovery Method"]:
        normalized["Website Discovery Method"] = "Not Found"
    if normalized["Website Verified"] == "":
        normalized["Website Verified"] = False
    return normalized


def dedupe_company_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    grouped: dict[str, dict[str, Any]] = {}
    input_rows = 0
    duplicate_rows = 0

    for raw_row in rows:
        row = normalize_company_row(raw_row)
        if not row.get("Company Name"):
            continue
        input_rows += 1
        key = company_dedupe_key(row)
        if key in grouped:
            duplicate_rows += 1
            grouped[key] = merge_company_rows(grouped[key], row)
        else:
            grouped[key] = row

    output_rows = []
    for row in grouped.values():
        row["Company ID"] = stable_company_id(row)
        output_rows.append(row)

    output_rows.sort(key=lambda row: str(row.get("Company Name", "")).lower())
    return output_rows, {
        "input_rows": input_rows,
        "duplicate_rows": duplicate_rows,
        "output_rows": len(output_rows),
    }


def merge_company_rows(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    preferred = best_row(left, right)
    merged = dict(preferred)
    other = right if preferred is left else left

    for column in OUTPUT_COLUMNS:
        if column == "Notes":
            merged[column] = combine_unique_notes(left.get(column, ""), right.get(column, ""))
        elif column == "Confidence":
            merged[column] = best_confidence(left.get(column), right.get(column))
        elif not merged.get(column) and other.get(column):
            merged[column] = other[column]

    return normalize_company_row(merged)


def best_row(left: dict[str, Any], right: dict[str, Any]) -> dict[str, Any]:
    left_score = row_quality_score(left)
    right_score = row_quality_score(right)
    return right if right_score > left_score else left


def row_quality_score(row: dict[str, Any]) -> tuple[int, int, int, int, str]:
    return (
        1 if row.get("Job Board URL") else 0,
        1 if row.get("Careers Page URL") else 0,
        1 if row.get("Official Website") else 0,
        normalize_confidence(row.get("Confidence")),
        str(row.get("Last Checked") or ""),
    )


def best_confidence(left: object, right: object) -> object:
    return right if normalize_confidence(right) > normalize_confidence(left) else left


def combine_unique_notes(*notes_values: object) -> str:
    notes: list[str] = []
    for value in notes_values:
        for note in str(value or "").split(";"):
            cleaned = note.strip()
            if cleaned and cleaned not in notes:
                notes.append(cleaned)
    return "; ".join(notes)


def company_dedupe_key(row: dict[str, Any]) -> str:
    domain = normalized_domain(str(row.get("Official Website") or row.get("Known Website") or ""))
    if domain:
        return f"domain:{domain}"
    return f"name-state:{slug(str(row.get('Company Name') or ''))}:{slug(str(row.get('State') or ''))}"


def stable_company_id(row: dict[str, Any]) -> str:
    domain = normalized_domain(str(row.get("Official Website") or row.get("Known Website") or ""))
    if domain:
        return f"company-{slug(domain)}"
    return f"company-{slug(str(row.get('Company Name') or ''))}-{slug(str(row.get('State') or ''))}"


def normalized_domain(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url if "://" in url else f"https://{url}")
    host = parsed.netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def slug(value: str) -> str:
    cleaned = []
    last_dash = False
    for char in value.lower():
        if char.isalnum():
            cleaned.append(char)
            last_dash = False
        elif not last_dash:
            cleaned.append("-")
            last_dash = True
    return "".join(cleaned).strip("-") or "unknown"


def create_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Companies"
    sheet.append(INPUT_COLUMNS)
    sheet.append(["BECU", "Tukwila", "WA", "https://www.becu.org", ""])
    sheet.append(["WECU", "Bellingham", "WA", "https://www.wecu.com", ""])
    sheet.append(["Canvas Credit Union", "Lone Tree", "CO", "", ""])
    sheet.append(["Ent Credit Union", "Colorado Springs", "CO", "", ""])
    sheet.append(["FirstBank", "Lakewood", "CO", "", ""])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = max(max_length + 2, 14)
    workbook.save(path)
